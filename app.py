# -*- coding: utf-8 -*-

import io
from flask import Flask, request, jsonify, send_file

# Se añade load_workbook para poder leer archivos Excel existentes
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule, DataBarRule
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.cell import MergedCell

# --- Inicialización de la Aplicación Flask ---
app = Flask(__name__)

# --- FUNCIONES DE AYUDA PARA CREAR EXCEL (SIN CAMBIOS) ---
def apply_styles_to_cell(cell, style_data):
    if not style_data or not isinstance(style_data, dict): return
    if 'font' in style_data: cell.font = Font(**style_data['font'])
    if 'fill' in style_data:
        if 'pattern' in style_data['fill']: style_data['fill']['fill_type'] = style_data['fill'].pop('pattern')
        cell.fill = PatternFill(**style_data['fill'])
    if 'border' in style_data:
        cell.border = Border(left=Side(**style_data['border'].get('left', {})), right=Side(**style_data['border'].get('right', {})), top=Side(**style_data['border'].get('top', {})), bottom=Side(**style_data['border'].get('bottom', {})))
    if 'alignment' in style_data: cell.alignment = Alignment(**style_data['alignment'])
    if 'numFmt' in style_data: cell.number_format = style_data['numFmt']

def create_chart_from_spec(worksheet, chart_spec):
    chart_type = chart_spec.get('type', 'bar').lower()
    if chart_type in ['bar', 'col']: chart = BarChart(); chart.type = chart_type
    elif chart_type == 'line': chart = LineChart()
    elif chart_type == 'pie': chart = PieChart()
    else: return
    chart.title = chart_spec.get('title', 'Gráfico sin Título')
    chart.style = chart_spec.get('style', 10)
    if 'y_axis_title' in chart_spec: chart.y_axis.title = chart_spec['y_axis_title']
    if 'x_axis_title' in chart_spec: chart.x_axis.title = chart_spec['x_axis_title']
    data = Reference(worksheet, range_string=chart_spec['data_range'])
    cats = Reference(worksheet, range_string=chart_spec['category_range'])
    chart.add_data(data, titles_from_data=chart_spec.get('titles_from_data', True))
    chart.set_categories(cats)
    worksheet.add_chart(chart, chart_spec.get('position', 'E1'))

# --- ENDPOINT PARA CREAR EXCEL DESDE JSON (SIN CAMBIOS) ---
@app.route('/create-excel', methods=['POST'])
def create_excel():
    # ... (Este endpoint ya estaba corregido y no necesita cambios)
    try:
        json_data = request.get_json()
        if not json_data or 'analysisData' not in json_data:
            return jsonify({"error": "El JSON no es válido o no contiene 'analysisData'."}), 400

        analysis_data = json_data.get('analysisData', [])
        conditional_rules = json_data.get('conditionalFormattingRules', [])
        chart_specs = json_data.get('charts', [])
        merge_cells_list = json_data.get('mergeCells', [])

        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Generado"

        for row_details in analysis_data:
            for cell_data in row_details:
                if cell_data and 'address' in cell_data:
                    cell = ws[cell_data['address']]
                    cell.value = cell_data.get('value')
                    apply_styles_to_cell(cell, cell_data.get('style'))

        for cell_range in merge_cells_list:
            ws.merge_cells(cell_range)

        for i, rule_info in enumerate(conditional_rules):
            try:
                style = rule_info.get('style', {})
                dxf = DifferentialStyle(font=Font(**style.get('font', {})), fill=PatternFill(**style.get('fill', {})))
                rule_params = {'type': rule_info['type'], 'dxf': dxf}
                if 'operator' in rule_info:
                    rule_params['operator'] = rule_info['operator']
                
                if rule_info['type'] == 'containsText':
                    if 'formulae' in rule_info and rule_info['formulae']:
                        rule_params['text'] = rule_info['formulae'][0]
                elif 'formulae' in rule_info:
                    cleaned_formulae = []
                    for f_val in rule_info['formulae']:
                        if f_val is not None and str(f_val).strip() != '':
                           cleaned_formulae.append(float(f_val))
                    if cleaned_formulae:
                        rule_params['formula'] = cleaned_formulae

                if rule_info['type'] == 'dataBar':
                    rule = DataBarRule(start_type='min', end_type='max', color=rule_info.get("color", "638EC6"))
                else:
                    if 'formula' in rule_params or 'text' in rule_params:
                        rule = Rule(**rule_params)
                        ws.conditional_formatting.add(rule_info['ref'], rule)
            except (ValueError, TypeError) as e:
                print(f"ADVERTENCIA: Se omitió la regla de formato condicional #{i} por datos inválidos.")
                print(f"   --- Regla problemática: {rule_info}")
                print(f"   --- Error específico: {e}")
                pass

        for spec in chart_specs:
            create_chart_from_spec(ws, spec)

        column_widths = {}
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell): continue
                try: length = len(str(cell.value))
                except: length = 0
                if cell.column not in column_widths or length > column_widths[cell.column]:
                    column_widths[cell.column] = length
        
        for col_idx, max_length in column_widths.items():
            column_letter = get_column_letter(col_idx)
            adjusted_width = (max_length + 2) if max_length < 50 else 50
            ws.column_dimensions[column_letter].width = adjusted_width

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return send_file(buffer, as_attachment=True, download_name='reporte_generado.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        print(f"Error en /create-excel: {e}")
        return jsonify({"error": f"Error interno del servidor: {str(e)}"}), 500


# === FUNCIÓN DE ANÁLISIS CORREGIDA ===
def extract_styles_from_cell(cell):
    """
    Extrae los estilos de una celda y los devuelve en un diccionario.
    Esta versión es más robusta y no usa el atributo '.is_default' que causaba el error.
    """
    style_data = {}
    if not cell.has_style:
        return style_data

    # Fuente (Font)
    font_data = {}
    if cell.font:
        if cell.font.name: font_data['name'] = cell.font.name
        if cell.font.sz: font_data['sz'] = cell.font.sz
        if cell.font.bold: font_data['bold'] = cell.font.bold
        if cell.font.italic: font_data['italic'] = cell.font.italic
        if cell.font.color and cell.font.color.rgb:
            font_data['color'] = cell.font.color.rgb
    if font_data:
        style_data['font'] = font_data

    # Relleno (Fill)
    fill_data = {}
    if cell.fill and cell.fill.fill_type:
        fill_data['pattern'] = cell.fill.fill_type
        if cell.fill.start_color and cell.fill.start_color.rgb:
            fill_data['start_color'] = cell.fill.start_color.rgb
        if cell.fill.end_color and cell.fill.end_color.rgb:
            fill_data['end_color'] = cell.fill.end_color.rgb
    if fill_data:
        style_data['fill'] = fill_data
        
    # Bordes (Border)
    border_data = {}
    if cell.border:
        def get_side_style(side):
            if side and side.style:
                return {'style': side.style, 'color': side.color.rgb if side.color else None}
            return None
        
        left, right, top, bottom = get_side_style(cell.border.left), get_side_style(cell.border.right), get_side_style(cell.border.top), get_side_style(cell.border.bottom)
        if left: border_data['left'] = left
        if right: border_data['right'] = right
        if top: border_data['top'] = top
        if bottom: border_data['bottom'] = bottom
    if border_data:
        style_data['border'] = border_data
        
    # Alineación (Alignment)
    alignment_data = {}
    if cell.alignment:
        if cell.alignment.horizontal: alignment_data['horizontal'] = cell.alignment.horizontal
        if cell.alignment.vertical: alignment_data['vertical'] = cell.alignment.vertical
        if cell.alignment.wrap_text: alignment_data['wrap_text'] = cell.alignment.wrap_text
    if alignment_data:
        style_data['alignment'] = alignment_data
        
    # Formato de número
    if cell.number_format and cell.number_format != 'General':
        style_data['numFmt'] = cell.number_format
            
    return style_data

# --- ENDPOINT PARA ANALIZAR EXCEL (CON LA FUNCIÓN CORREGIDA) ---
@app.route('/parse-excel', methods=['POST'])
def parse_excel():
    if 'excel_file' not in request.files:
        return jsonify({"error": "No se encontró el archivo en la petición (se esperaba el campo 'excel_file')."}), 400
    file = request.files['excel_file']
    if file.filename == '':
        return jsonify({"error": "No se seleccionó ningún archivo."}), 400
    try:
        in_memory_file = io.BytesIO(file.read())
        wb = load_workbook(filename=in_memory_file, data_only=True)
        parsed_data = {'sheets': []}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_data = {
                'name': sheet_name,
                'data': [],
                'merged_cells': [str(merged_range) for merged_range in ws.merged_cells.ranges]
            }
            rows_data = []
            for row in ws.iter_rows():
                row_list = []
                for cell in row:
                    cell_info = {
                        'address': cell.coordinate,
                        'value': cell.value
                    }
                    if isinstance(cell, MergedCell):
                        cell_info['is_merged_part'] = True
                    else:
                        # Usando la función corregida
                        cell_info['style'] = extract_styles_from_cell(cell)
                    row_list.append(cell_info)
                rows_data.append(row_list)
            sheet_data['data'] = rows_data
            parsed_data['sheets'].append(sheet_data)
        return jsonify(parsed_data)
    except Exception as e:
        print(f"Error en /parse-excel: {e}")
        return jsonify({"error": f"Error interno al procesar el archivo Excel: {str(e)}"}), 500

# --- Punto de Entrada de la Aplicación ---
if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
